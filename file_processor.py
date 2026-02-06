"""
File Processor for handling various file types for company policies
Supports PDF, DOCX, TXT, MD, and other text-based formats
"""
import os
from pathlib import Path
from typing import List, Optional
import io


class FileProcessor:
    """Process various file types to extract text content"""
    
    @staticmethod
    def extract_text_from_file(file_path: str) -> str:
        """
        Extract text from various file types
        
        Args:
            file_path: Path to the file
            
        Returns:
            Extracted text content
        """
        file_ext = Path(file_path).suffix.lower()
        
        try:
            if file_ext == '.pdf':
                return FileProcessor._extract_from_pdf(file_path)
            elif file_ext in ['.docx', '.doc']:
                return FileProcessor._extract_from_docx(file_path)
            elif file_ext in ['.txt', '.md', '.markdown']:
                return FileProcessor._extract_from_text(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return FileProcessor._extract_from_excel(file_path)
            elif file_ext in ['.csv']:
                return FileProcessor._extract_from_csv(file_path)
            else:
                # Try as text file
                return FileProcessor._extract_from_text(file_path)
        except Exception as e:
            return f"[Error reading file {file_path}: {str(e)}]"
    
    @staticmethod
    def _extract_from_pdf(file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            import PyPDF2
            text = []
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text.append(page.extract_text())
            return '\n\n'.join(text)
        except ImportError:
            return "[PDF extraction requires PyPDF2. Install with: pip install PyPDF2]"
        except Exception as e:
            return f"[PDF extraction error: {str(e)}]"
    
    @staticmethod
    def _extract_from_docx(file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            from docx import Document
            doc = Document(file_path)
            text = []
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
            return '\n\n'.join(text)
        except ImportError:
            return "[DOCX extraction requires python-docx. Install with: pip install python-docx]"
        except Exception as e:
            return f"[DOCX extraction error: {str(e)}]"
    
    @staticmethod
    def _extract_from_text(file_path: str) -> str:
        """Extract text from plain text file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as file:
                        return file.read()
                except UnicodeDecodeError:
                    continue
            # If all encodings fail, read as binary and decode with errors='replace'
            with open(file_path, 'rb') as file:
                return file.read().decode('utf-8', errors='replace')
        except Exception as e:
            return f"[Text extraction error: {str(e)}]"
    
    @staticmethod
    def _extract_from_excel(file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path)
            text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"\n--- Sheet: {sheet_name} ---\n")
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                    if row_text.strip():
                        text.append(row_text)
            return '\n'.join(text)
        except ImportError:
            return "[Excel extraction requires openpyxl. Install with: pip install openpyxl]"
        except Exception as e:
            return f"[Excel extraction error: {str(e)}]"
    
    @staticmethod
    def _extract_from_csv(file_path: str) -> str:
        """Extract text from CSV file"""
        try:
            import csv
            text = []
            with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    text.append(' | '.join(row))
            return '\n'.join(text)
        except Exception as e:
            return f"[CSV extraction error: {str(e)}]"
    
    @staticmethod
    def process_multiple_files(file_paths: List[str]) -> str:
        """
        Process multiple files and combine their content
        
        Args:
            file_paths: List of file paths
            
        Returns:
            Combined text content from all files
        """
        combined_text = []
        
        for file_path in file_paths:
            if os.path.exists(file_path):
                file_name = Path(file_path).name
                combined_text.append(f"\n{'='*70}\n")
                combined_text.append(f"File: {file_name}\n")
                combined_text.append(f"{'='*70}\n\n")
                content = FileProcessor.extract_text_from_file(file_path)
                combined_text.append(content)
                combined_text.append("\n")
        
        return '\n'.join(combined_text)

